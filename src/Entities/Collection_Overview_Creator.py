import collections
import csv
import os
from pathlib import Path
from PyPDF2 import PdfMerger
from openpyxl import load_workbook
from openpyxl_image_loader import SheetImageLoader
from pdfkit import from_string
from src.Entities.Collection_Entry import Collection_Entry
from src.Entities.template_factory import Template_Factory

class Collection_Overview_Creator:

    @staticmethod
    def create_collection_overview(file_path_list):

        root_path = Path(__file__).parents[2]

        website_individual_html_folder_path = root_path.joinpath("resources", "website", "individual_htmls")

        print_full_page_html_folder_path = root_path.joinpath("resources", "print", "full_page")
        print_image_page_html_folder_path = root_path.joinpath("resources", "print", "image_page")

        final_pdf_file_path = root_path.joinpath("resources", "Sammlungs_Katalog_2022.pdf")
        image_folder_path = root_path.joinpath("resources", "images")

        missing_images_csv_path = root_path.joinpath("resources", "cleaning_info", "missing_images.csv")
        artists_csv_path = root_path.joinpath("resources", "cleaning_info", "artists.csv")
        long_title_path = root_path.joinpath("resources", "cleaning_info", "long_title.csv")

        art_count = 1
        individual_artworks_list = []
        artist_artwork_dic = collections.defaultdict(list)

        no_image_list = []
        artists_dic = {}
        long_title_list = []

        for file_index, file_path in enumerate(file_path_list):

            file_name = os.path.basename(file_path)

            wb = load_workbook(file_path, data_only=True)
            sheet = wb['KUNSTW_20_B']

            active_rows = wb.active
            max_rows = active_rows.max_row

            # Put your sheet in the loader
            image_loader = SheetImageLoader(sheet)

            row_count = 0
            keep_going = True

            artist_information = []

            for row in sheet.iter_rows():

                print_string = f"File {file_index + 1} / {len(file_path_list)} - Row {row_count + 1} / {max_rows}"
                print(print_string)

                if row_count > 0 and keep_going:
                    cell_count = 0

                    hochformat = 'O' + str(row_count + 1)
                    querformat = 'P' + str(row_count + 1)
                    quadratisch = 'Q' + str(row_count + 1)

                    image_file_path = str(image_folder_path) + "\\" + str(art_count) + ".jpg"

                    try:
                        hoch_image = image_loader.get(hochformat)
                        if hoch_image.mode != 'RGB':
                            hoch_image = hoch_image.convert('RGB')
                        hoch_image.save(image_file_path)
                        image_format = "hoch"
                        image_path = image_file_path
                    except Exception as e:
                        error_message = str(e)
                        if "doesn't contain an image" not in error_message:
                            print(f"Hoch error - {e}")
                        try:
                            quer_image = image_loader.get(querformat)
                            if quer_image.mode != 'RGB':
                                quer_image = quer_image.convert('RGB')
                            quer_image.save(image_file_path)
                            image_format = "quer"
                            image_path = image_file_path
                        except Exception as e:
                            error_message = str(e)
                            if "doesn't contain an image" not in error_message:
                                print(f"Quer error - {e}")
                            try:
                                quadratisch_image = image_loader.get(quadratisch)
                                if quadratisch_image.mode != 'RGB':
                                    quadratisch_image = quadratisch_image.convert('RGB')
                                quadratisch_image.save(image_file_path)
                                image_format = "quad"
                                image_path = image_file_path
                            except Exception as e:
                                error_message = str(e)
                                if "doesn't contain an image" not in error_message:
                                    print(f"Quad error - {e}")

                                print(f"Row {row_count + 1}: No image")
                                no_image_list.append((row_count + 1, file_name))

                                image_format = "no"
                                image_path = "no"

                    for cell in row:

                        if keep_going:

                            if cell_count == 0:
                                if cell.value == None:
                                    keep_going = False
                                else:
                                    id = int(cell.value)

                            elif cell_count == 1:
                                location = cell.value
                            elif cell_count == 2:
                                ken_1 = cell.value
                            elif cell_count == 3:
                                ken_2 = cell.value
                            elif cell_count == 4:
                                count = cell.value
                            elif cell_count == 5:
                                series = cell.value
                            elif cell_count == 6:
                                artist = cell.value
                            elif cell_count == 7:
                                country = cell.value
                            elif cell_count == 8:
                                title = cell.value
                            elif cell_count == 9:
                                technique_count = cell.value
                            elif cell_count == 10:
                                material = cell.value
                            elif cell_count == 11:
                                production = cell.value
                            elif cell_count == 12:
                                frame = cell.value
                            elif cell_count == 13:
                                size = cell.value

                            elif cell_count == 19:
                                edition = cell.value
                            elif cell_count == 20:
                                edition_pp = cell.value
                            elif cell_count == 21:
                                edition_total = cell.value
                            elif cell_count == 22:
                                number = cell.value
                            elif cell_count == 23:
                                signature = cell.value
                            elif cell_count == 24:
                                certificate = cell.value
                            elif cell_count == 25:
                                price = cell.value
                            elif cell_count == 27:
                                seller = cell.value
                            elif cell_count == 28:
                                sale_month = cell.value
                            elif cell_count == 29:
                                sale_year = cell.value
                            elif cell_count == 31:
                                miscellaneous = cell.value

                            cell_count += 1

                    if keep_going:
                        newCollectionEntry = Collection_Entry(id, location, ken_1, ken_2, count, series, artist,
                                                              country, title, technique_count, material, production,
                                                              frame, size, image_format, image_path, edition,
                                                              edition_pp, edition_total, number, signature, certificate,
                                                              price,
                                                              seller, sale_month, sale_year, miscellaneous)
                        individual_artworks_list.append(newCollectionEntry)

                        artist_artwork_dic[artist].append((art_count, title, image_path))

                        if artist not in artists_dic:
                            artists_dic[artist] = {}
                        if "file" not in artists_dic[artist]:
                            artists_dic[artist]["file"] = file_name
                        if "rows" not in artists_dic[artist]:
                            artists_dic[artist]["rows"] = []

                        artists_dic[artist]["rows"].append(row_count + 1)

                        if title and len(str(title)) > 10:
                            long_title_list.append((title, row_count + 1, file_name))

                        art_count += 1

                row_count += 1


        with open(missing_images_csv_path, 'w', newline='', encoding='utf-8') as missing_images_csv_file:
            writer = csv.DictWriter(missing_images_csv_file, fieldnames=["row", "file"])
            writer.writeheader()
            for row, file in no_image_list:
                csv_dic = {}
                csv_dic["row"] = str(row)
                csv_dic["file"] = str(file)
                writer.writerow(csv_dic)

        with open(artists_csv_path, 'w', newline='', encoding='utf-8') as artist_csv_file:
            writer = csv.DictWriter(artist_csv_file, fieldnames=["artist", "file", "rows"])
            writer.writeheader()
            for artist, info_dic in artists_dic.items():
                csv_dic = {}
                csv_dic["artist"] = str(artist)
                csv_dic["file"] = str(info_dic["file"])
                csv_dic["rows"] = "_".join([str(x) for x in info_dic["rows"]])
                writer.writerow(csv_dic)

        with open(long_title_path, 'w', newline='', encoding='utf-8') as long_title_csv_file:
            writer = csv.DictWriter(long_title_csv_file, fieldnames=["title", "row", "file"])
            writer.writeheader()
            for title, row, file in long_title_list:
                csv_dic = {}
                csv_dic["title"] = str(title)
                csv_dic["row"] = str(row)
                csv_dic["file"] = str(file)
                writer.writerow(csv_dic)

        for count, collectionEntry in enumerate(individual_artworks_list):
            print_string = str(count + 1) + "/" + str(len(individual_artworks_list))
            print(print_string)

            print_full_page_file_name = str(count + 1) + ".html"
            print_full_page_file_path = str(print_full_page_html_folder_path) + "\\" + print_full_page_file_name

            print_full_page_html_string = Template_Factory.create_print_full_page(collectionEntry=collectionEntry)
            with open(print_full_page_file_path, "w", encoding="utf-8") as print_full_page_html_file:
                print_full_page_html_file.write(print_full_page_html_string)
            print_full_page_html_file.close()


            print_image_page_file_name = str(count + 1) + ".html"
            print_image_page_file_path = str(print_image_page_html_folder_path) + "\\" + print_image_page_file_name

            print_image_page_html_string = Template_Factory.create_print_image_page(collectionEntry=collectionEntry)
            with open(print_image_page_file_path, "w", encoding="utf-8") as print_image_page_html_file:
                print_image_page_html_file.write(print_image_page_html_string)
            print_image_page_html_file.close()
